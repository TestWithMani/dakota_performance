"""Excel logger for rich performance measurement outputs."""

from __future__ import annotations

import math
import statistics
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config

HEADERS = [
    "Row Type",
    "Tab",
    "Sample #",
    "Time (s)",
    "Min (s)",
    "Max (s)",
    "Performance Benchmark (s)",
    "Result",
    "Browser",
    "Recorded At",
    "Platform",
    "Notes",
]
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PASS_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="E8EEF8")
FAIL_RESULT_FONT = Font(color="C00000", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _ensure_workbook(file_path: Path) -> None:
    """Create workbook with headers when it does not exist."""
    if file_path.exists():
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dakota Marketplace Performance"
    sheet.append(HEADERS)
    _style_header(sheet)
    _set_sheet_layout(sheet)
    workbook.save(file_path)


def log_performance_results(
    tab_name: str,
    execution_times: Iterable[float],
    average_time: float,
    sla_seconds: float,
    browser: str = "Unknown",
    platform: str = "Unknown",
    os_version: str = "Unknown",
    file_name: str = config.EXCEL_FILE_NAME,
) -> Path:
    """Write detailed iteration rows and one run summary row into Excel."""
    file_path = Path(file_name).resolve()
    _ensure_workbook(file_path)

    workbook = load_workbook(file_path)
    sheet = workbook.active
    if sheet.title != "Dakota Marketplace Performance":
        sheet.title = "Dakota Marketplace Performance"
    _ensure_headers(sheet)
    _style_header(sheet)
    _set_sheet_layout(sheet)

    timings = [round(value, 3) for value in execution_times]
    run_id = str(uuid.uuid4())
    now_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    min_time = round(min(timings), 3)
    max_time = round(max(timings), 3)
    std_dev = round(statistics.pstdev(timings), 3) if len(timings) > 1 else 0.0
    result = "PASS" if average_time <= sla_seconds else "FAIL"

    previous_avg = _get_previous_run_average(sheet, tab_name)
    previous_run_note = "N/A (first or unavailable)"
    if previous_avg is not None and previous_avg > 0:
        change_pct = round(((average_time - previous_avg) / previous_avg) * 100, 2)
        trend = "Improved" if change_pct < 0 else "Degraded" if change_pct > 0 else "No change"
        previous_run_note = f"{trend} ({change_pct:+.2f}%)"

    for idx, duration in enumerate(timings, start=1):
        row = [
            "Iteration",
            tab_name,
            idx,
            duration,
            "",
            "",
            "",
            "",
            browser,
            now_utc,
            platform,
            "",
        ]
        sheet.append(row)
        # Iteration rows are informational only; SLA/result is evaluated at run-summary level.
        _style_data_row(sheet, sheet.max_row, is_summary=False, is_pass=True)

    summary_notes = (
        f"Samples={len(timings)} | Attempts={len(timings)} | skipped network=0 "
        f"| skipped marker=0 | prev run: {previous_run_note}"
    )
    summary_row = [
        "Run summary",
        tab_name,
        "",
        round(average_time, 3),
        min_time,
        max_time,
        round(sla_seconds, 3),
        result,
        browser,
        now_utc,
        platform,
        summary_notes,
    ]
    sheet.append(summary_row)
    _style_data_row(sheet, sheet.max_row, is_summary=True, is_pass=result == "PASS")

    _set_sheet_layout(sheet)
    workbook.save(file_path)
    return file_path


def _get_previous_run_average(sheet, tab_name: str) -> float | None:
    """Return previous run summary average time for this tab, if present."""
    if sheet.max_row < 2:
        return None

    # Traverse from bottom to top and skip the current run-in-progress rows.
    for row_idx in range(sheet.max_row, 1, -1):
        row_type = sheet.cell(row=row_idx, column=1).value
        row_tab = sheet.cell(row=row_idx, column=2).value
        time_value = sheet.cell(row=row_idx, column=4).value

        if row_type == "Run summary" and row_tab == tab_name:
            try:
                parsed = float(time_value)
                if not math.isnan(parsed):
                    return round(parsed, 3)
            except (TypeError, ValueError):
                return None
    return None


def _style_header(sheet) -> None:
    """Apply strong table header styling."""
    for col_index, _ in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_row(sheet, row_number: int, is_summary: bool, is_pass: bool) -> None:
    """Apply row styling by type and result."""
    if is_summary:
        row_fill = SUMMARY_FILL
    else:
        row_fill = PASS_FILL if is_pass else FAIL_FILL

    for col_index in range(1, len(HEADERS) + 1):
        cell = sheet.cell(row=row_number, column=col_index)
        cell.fill = row_fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_index == 12:  # Notes
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if is_summary:
            cell.font = Font(bold=True)

    # Always make FAIL result visibly red in the Result column.
    result_cell = sheet.cell(row=row_number, column=8)
    if str(result_cell.value).strip().upper() == "FAIL":
        result_cell.font = FAIL_RESULT_FONT


def _set_sheet_layout(sheet) -> None:
    """Set widths, freeze header, and apply table usability settings."""
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{sheet.max_row}"
    sheet.sheet_view.showGridLines = True

    column_widths = {
        "A": 12,
        "B": 26,
        "C": 10,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 22,
        "H": 10,
        "I": 22,
        "J": 16,
        "K": 16,
        "L": 72,
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width


def _ensure_headers(sheet) -> None:
    """Ensure row 1 always matches the current schema."""
    if sheet.max_row == 0:
        sheet.append(HEADERS)
        return

    for col_index, header in enumerate(HEADERS, start=1):
        sheet.cell(row=1, column=col_index, value=header)
